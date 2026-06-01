"""
Export endpoints — XLSX downloads for dashboard data.

GET /export/users           →  /api/dashboard/export/export/users
GET /export/users/{user_id} →  /api/dashboard/export/export/users/{user_id}

Global export: all users matching filters → Users_Summary + Meta sheets.
Single-user export: deep per-user workbook → Summary + Trend_30d +
                    Top_Motion_Types + Recent_Sessions + Recent_Activity + Meta.
"""

import io
import json
from datetime import date, datetime, timedelta, timezone
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text

from ...auth.models import User
from ...common.dependencies import get_current_firm_user
from ...auth.database import UserAsyncSessionLocal
from ...chatbot.database import AsyncSessionLocal
from ._shared import DateRangeParams, _ACTION_LABELS, _build_activity_detail, _MOTION_DISPLAY_NAMES

router = APIRouter(tags=["exports"])

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_KEY_FONT     = Font(bold=True)


def _fmt_dt(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M UTC")


def _apply_header_row(ws, headers: list[str], col_widths: list[int]) -> None:
    ws.append(headers)
    ws.row_dimensions[1].height = 20
    for i, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]
    ws.freeze_panes = "A2"


def _write_meta_sheet(wb: openpyxl.Workbook, rows: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet("Meta")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 52
    for key, val in rows:
        ws.append([key, val])
        ws.cell(row=ws.max_row, column=1).font = _KEY_FONT


# ---------------------------------------------------------------------------
# GET /export/users  — Global users export
# ---------------------------------------------------------------------------

@router.get("/export/users")
async def export_users_xlsx(
    dr: DateRangeParams = Depends(),
    _user: User = Depends(get_current_firm_user),
    search: Optional[str] = Query(None, description="Filter by name or email (case-insensitive)"),
    sort_by: str = Query("last_active", pattern="^(last_active|cases_count|motions_drafted|created_at)$"),
    sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
    activity_status: Optional[str] = Query(None, pattern="^(active|inactive|new)$"),
    source: Optional[str] = Query(None, pattern="^(manual|ecf|gdrive|courtdrive)$"),
):
    """
    Download all users as an XLSX workbook (Users_Summary + Meta sheets).

    Accepts the same search / sort params as GET /analytics/users, plus
    activity_status and source extension filters.
    """
    range_params = {"start": dr.start, "end": dr.end}
    _dt_min = datetime.min.replace(tzinfo=timezone.utc)

    # ── Step 1: user_db — all matching users ─────────────────────────────
    async with UserAsyncSessionLocal() as user_db:
        if search:
            id_result = await user_db.execute(
                text("""
                    SELECT id, email, first_name, last_name, created_at FROM users
                    WHERE is_active = true
                      AND (LOWER(email) LIKE :pat
                           OR LOWER(COALESCE(first_name, '') || ' ' || COALESCE(last_name, '')) LIKE :pat)
                """),
                {"pat": f"%{search.lower()}%"},
            )
        else:
            id_result = await user_db.execute(
                text("SELECT id, email, first_name, last_name, created_at FROM users WHERE is_active = true")
            )
        user_rows = id_result.fetchall()

    all_user_ids: list[str] = [r.id for r in user_rows]
    user_info_map = {r.id: r for r in user_rows}
    created_at_map: dict[str, datetime] = {r.id: r.created_at for r in user_rows}

    last_active_map: dict[str, datetime] = {}
    cases_map: dict[str, int] = {}
    motions_map: dict[str, dict] = {}
    engagement_map: dict[str, dict] = {}
    top_motions_map: dict[str, list[str]] = {}
    active_user_ids_set: set[str] = set()
    source_user_ids: Optional[set[str]] = None
    sort_key_map: dict = {}
    default_sort: object = _dt_min

    if all_user_ids:
        async with AsyncSessionLocal() as chat_db:
            # ── Sort key aggregate ────────────────────────────────────────
            if sort_by == "last_active":
                sk_rows = await chat_db.execute(
                    text("""
                        SELECT user_id, MAX(created_at) AS sort_val
                        FROM user_activity_logs WHERE user_id = ANY(:ids)
                        GROUP BY user_id
                    """),
                    {"ids": all_user_ids},
                )
                sort_key_map = {r.user_id: r.sort_val for r in sk_rows}
                default_sort = _dt_min
            elif sort_by == "cases_count":
                sk_rows = await chat_db.execute(
                    text("""
                        WITH lp AS (
                            SELECT DISTINCT ON (session_id) session_id, petition_status
                            FROM pdf_documents ORDER BY session_id, uploaded_at DESC
                        )
                        SELECT s.user_id, COUNT(DISTINCT s.id) AS sort_val
                        FROM sessions s
                        INNER JOIN lp ON lp.session_id = s.id
                        INNER JOIN chat_threads ct ON ct.session_id = s.id AND ct.is_active = true
                        WHERE s.user_id = ANY(:ids) AND s.is_active = true
                          AND s.created_at >= :start AND s.created_at <= :end
                          AND (lp.petition_status IN ('working', 'accepted') OR lp.petition_status IS NULL)
                        GROUP BY s.user_id
                    """),
                    {"ids": all_user_ids, **range_params},
                )
                sort_key_map = {r.user_id: int(r.sort_val) for r in sk_rows}
                default_sort = 0
            elif sort_by == "motions_drafted":
                sk_rows = await chat_db.execute(
                    text("""
                        SELECT s.user_id, COUNT(m.id) AS sort_val
                        FROM motion_draft_logs m
                        INNER JOIN sessions s ON s.id = m.session_id
                        WHERE s.user_id = ANY(:ids)
                          AND m.created_at >= :start AND m.created_at <= :end
                        GROUP BY s.user_id
                    """),
                    {"ids": all_user_ids, **range_params},
                )
                sort_key_map = {r.user_id: int(r.sort_val) for r in sk_rows}
                default_sort = 0
            else:  # created_at
                sort_key_map = created_at_map
                default_sort = _dt_min

            # ── Last active ───────────────────────────────────────────────
            la_rows = await chat_db.execute(
                text("""
                    SELECT user_id, MAX(created_at) AS last_active_at
                    FROM user_activity_logs WHERE user_id = ANY(:ids)
                    GROUP BY user_id
                """),
                {"ids": all_user_ids},
            )
            last_active_map = {r.user_id: r.last_active_at for r in la_rows}

            # ── Cases count ───────────────────────────────────────────────
            c_rows = await chat_db.execute(
                text("""
                    WITH lp AS (
                        SELECT DISTINCT ON (session_id) session_id, petition_status
                        FROM pdf_documents ORDER BY session_id, uploaded_at DESC
                    )
                    SELECT s.user_id, COUNT(DISTINCT s.id) AS cases_count
                    FROM sessions s
                    INNER JOIN lp ON lp.session_id = s.id
                    INNER JOIN chat_threads ct ON ct.session_id = s.id AND ct.is_active = true
                    WHERE s.user_id = ANY(:ids) AND s.is_active = true
                      AND s.created_at >= :start AND s.created_at <= :end
                      AND (lp.petition_status IN ('working', 'accepted') OR lp.petition_status IS NULL)
                    GROUP BY s.user_id
                """),
                {"ids": all_user_ids, **range_params},
            )
            cases_map = {r.user_id: int(r.cases_count) for r in c_rows}

            # ── Source filter ─────────────────────────────────────────────
            if source:
                src_rows = await chat_db.execute(
                    text("""
                        SELECT DISTINCT s.user_id
                        FROM sessions s
                        INNER JOIN pdf_documents p ON p.session_id = s.id
                        WHERE s.user_id = ANY(:ids) AND p.source = :source
                    """),
                    {"ids": all_user_ids, "source": source},
                )
                source_user_ids = {r.user_id for r in src_rows}

            # ── Motions: started, completed, avg time ─────────────────────
            m_rows = await chat_db.execute(
                text("""
                    SELECT
                        s.user_id,
                        COUNT(m.id)                                               AS motions_started,
                        COUNT(m.id) FILTER (WHERE m.status = 'completed')         AS motions_completed,
                        AVG(
                            CASE WHEN m.status = 'completed' AND m.completed_at IS NOT NULL
                            THEN EXTRACT(EPOCH FROM (m.completed_at - m.created_at)) END
                        )                                                          AS avg_draft_time_seconds
                    FROM motion_draft_logs m
                    INNER JOIN sessions s ON s.id = m.session_id
                    WHERE s.user_id = ANY(:ids)
                      AND m.created_at >= :start AND m.created_at <= :end
                    GROUP BY s.user_id
                """),
                {"ids": all_user_ids, **range_params},
            )
            for r in m_rows:
                started  = int(r.motions_started)
                done     = int(r.motions_completed)
                rate     = round(done * 100.0 / started, 1) if started > 0 else None
                avg_s    = float(r.avg_draft_time_seconds) if r.avg_draft_time_seconds is not None else None
                motions_map[r.user_id] = {
                    "started":      started,
                    "completed":    done,
                    "success_rate": rate,
                    "avg_seconds":  avg_s,
                }

            # ── Engagement: login count, active days, docs exported ───────
            eng_rows = await chat_db.execute(
                text("""
                    SELECT
                        user_id,
                        COUNT(*) FILTER (WHERE action = 'login')                              AS login_count,
                        COUNT(DISTINCT created_at::date)                                      AS active_days,
                        COUNT(*) FILTER (
                            WHERE action IN ('generate_document', 'download_motion')
                        )                                                                      AS docs_exported
                    FROM user_activity_logs
                    WHERE user_id = ANY(:ids)
                      AND created_at >= :start AND created_at <= :end
                    GROUP BY user_id
                """),
                {"ids": all_user_ids, **range_params},
            )
            for r in eng_rows:
                engagement_map[r.user_id] = {
                    "login_count":  int(r.login_count),
                    "active_days":  int(r.active_days),
                    "docs_exported": int(r.docs_exported),
                }
                if int(r.login_count) > 0 or int(r.active_days) > 0:
                    active_user_ids_set.add(r.user_id)

            # ── Sessions created in range ─────────────────────────────────
            sess_rows = await chat_db.execute(
                text("""
                    SELECT user_id, COUNT(*) AS sessions_created
                    FROM sessions
                    WHERE user_id = ANY(:ids)
                      AND created_at >= :start AND created_at <= :end
                    GROUP BY user_id
                """),
                {"ids": all_user_ids, **range_params},
            )
            for r in sess_rows:
                uid = r.user_id
                sc  = int(r.sessions_created)
                if uid in engagement_map:
                    engagement_map[uid]["sessions_created"] = sc
                else:
                    engagement_map[uid] = {
                        "login_count": 0, "active_days": 0,
                        "docs_exported": 0, "sessions_created": sc,
                    }

            # ── Top 3 motion types per user (completed, in range) ─────────
            tm_rows = await chat_db.execute(
                text("""
                    WITH ranked AS (
                        SELECT
                            s.user_id,
                            m.motion_type,
                            COUNT(*) AS cnt,
                            ROW_NUMBER() OVER (
                                PARTITION BY s.user_id ORDER BY COUNT(*) DESC
                            ) AS rn
                        FROM motion_draft_logs m
                        INNER JOIN sessions s ON s.id = m.session_id
                        WHERE s.user_id = ANY(:ids) AND m.status = 'completed'
                          AND m.created_at >= :start AND m.created_at <= :end
                        GROUP BY s.user_id, m.motion_type
                    )
                    SELECT user_id, motion_type FROM ranked WHERE rn <= 3
                    ORDER BY user_id, cnt DESC
                """),
                {"ids": all_user_ids, **range_params},
            )
            for r in tm_rows:
                top_motions_map.setdefault(r.user_id, []).append(r.motion_type)

    # ── Step 2: sort and apply extension filters ──────────────────────────
    reverse    = sort_dir == "desc"
    sorted_ids = sorted(
        all_user_ids,
        key=lambda uid: sort_key_map.get(uid) or default_sort,
        reverse=reverse,
    )

    if activity_status == "active":
        sorted_ids = [uid for uid in sorted_ids if uid in active_user_ids_set]
    elif activity_status == "inactive":
        sorted_ids = [uid for uid in sorted_ids if uid not in active_user_ids_set]
    elif activity_status == "new":
        new_ids = {r.id for r in user_rows if r.created_at and r.created_at >= dr.start}
        sorted_ids = [uid for uid in sorted_ids if uid in new_ids]

    if source_user_ids is not None:
        sorted_ids = [uid for uid in sorted_ids if uid in source_user_ids]

    # ── Step 3: build XLSX ────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users_Summary"

    headers = [
        "Row #", "User ID", "Name", "Email",
        "Joined At (UTC)", "Last Active At (UTC)",
        "Cases (in range)",
        "Motions Started (in range)", "Motions Completed (in range)",
        "Draft Success Rate (%)", "Avg Draft Time (min)",
        "Login Count", "Active Days", "Sessions Created",
        "Documents Exported",
        "Top Motion Type #1", "Top Motion Type #2", "Top Motion Type #3",
    ]
    col_widths = [6, 38, 28, 36, 22, 22, 18, 26, 28, 22, 22, 14, 13, 18, 20, 30, 30, 30]
    _apply_header_row(ws, headers, col_widths)

    for row_num, uid in enumerate(sorted_ids, start=1):
        u = user_info_map.get(uid)
        if not u:
            continue
        name     = " ".join(filter(None, [u.first_name, u.last_name])) or u.email
        m        = motions_map.get(uid, {})
        eng      = engagement_map.get(uid, {})
        avg_s    = m.get("avg_seconds")
        avg_min  = round(avg_s / 60, 1) if avg_s is not None else None
        tops     = top_motions_map.get(uid, [])
        rate     = m.get("success_rate")

        ws.append([
            row_num, uid, name, u.email,
            _fmt_dt(u.created_at),
            _fmt_dt(last_active_map.get(uid)),
            cases_map.get(uid, 0),
            m.get("started", 0),
            m.get("completed", 0),
            rate,
            avg_min,
            eng.get("login_count", 0),
            eng.get("active_days", 0),
            eng.get("sessions_created", 0),
            eng.get("docs_exported", 0),
            tops[0] if len(tops) > 0 else None,
            tops[1] if len(tops) > 1 else None,
            tops[2] if len(tops) > 2 else None,
        ])

    applied_filters: dict = {}
    if search:
        applied_filters["search"] = search
    if activity_status:
        applied_filters["activity_status"] = activity_status
    if source:
        applied_filters["source"] = source

    _write_meta_sheet(wb, [
        ("Export Generated At (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Export Type", "global_users"),
        ("Range Preset", dr.range),
        ("Range Start (UTC)", dr.start.strftime("%Y-%m-%d %H:%M UTC")),
        ("Range End (UTC)", dr.end.strftime("%Y-%m-%d %H:%M UTC")),
        ("Applied Search", search or ""),
        ("Applied Filters (JSON)", json.dumps(applied_filters)),
        ("Applied Sort (JSON)", json.dumps({"sort_by": sort_by, "sort_dir": sort_dir})),
        ("Total Rows Exported", str(len(sorted_ids))),
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    date_tag = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"users_{dr.range}_{date_tag}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# GET /export/users/{user_id}  — Single-user detail export
# ---------------------------------------------------------------------------

@router.get("/export/users/{user_id}")
async def export_single_user_xlsx(
    user_id: str,
    dr: DateRangeParams = Depends(),
    _user: User = Depends(get_current_firm_user),
    # Sessions scope
    sessions_page: int = Query(1, ge=1),
    sessions_page_size: int = Query(10, ge=1, le=50),
    sessions_search: Optional[str] = Query(None),
    sessions_source: Optional[str] = Query(None, pattern="^(manual|ecf|gdrive|courtdrive)$"),
    sessions_status: Optional[str] = Query(None, pattern="^(working|accepted|pending_acceptance|archived)$"),
    sessions_sort_by: str = Query("last_activity", pattern="^(case|source|status|motions|last_activity)$"),
    sessions_sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
    # Activity scope
    activity_page: int = Query(1, ge=1),
    activity_page_size: int = Query(10, ge=1, le=50),
    activity_search: Optional[str] = Query(None),
    activity_action: Optional[str] = Query(None),
    activity_sort_by: str = Query("occurred", pattern="^(action|duration|occurred)$"),
    activity_sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
):
    """
    Download a single user's analytics as a multi-sheet XLSX workbook.

    Sheets: Summary, Trend_30d, Top_Motion_Types, Recent_Sessions,
            Recent_Activity, Meta.
    """
    range_params = {"start": dr.start, "end": dr.end}

    # ── Resolve user identity ─────────────────────────────────────────────
    async with UserAsyncSessionLocal() as user_db:
        u_result = await user_db.execute(
            text("""
                SELECT id, email, first_name, last_name, created_at
                FROM users WHERE id = :uid AND is_active = true
            """),
            {"uid": user_id},
        )
        u = u_result.fetchone()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    display_name = " ".join(filter(None, [u.first_name, u.last_name])) or u.email

    # ── Fetch all behavioral data in one chat_db session ─────────────────
    summary: dict = {}
    trend_rows: list[tuple] = []
    top_motion_rows: list[tuple] = []
    session_rows: list = []
    activity_rows: list = []

    async with AsyncSessionLocal() as chat_db:
        # Last active (all time)
        la = await chat_db.execute(
            text("SELECT MAX(created_at) AS v FROM user_activity_logs WHERE user_id = :uid"),
            {"uid": user_id},
        )
        summary["last_active_at"] = la.scalar()

        # Engagement KPIs in range
        eng = await chat_db.execute(
            text("""
                SELECT
                    COUNT(*) FILTER (WHERE action = 'login')                         AS login_count,
                    COUNT(DISTINCT created_at::date)                                  AS active_days,
                    COUNT(*) FILTER (
                        WHERE action IN ('generate_document', 'download_motion')
                    )                                                                  AS docs_exported
                FROM user_activity_logs
                WHERE user_id = :uid AND created_at >= :start AND created_at <= :end
            """),
            {"uid": user_id, **range_params},
        )
        er = eng.fetchone()
        summary["login_count"]   = int(er.login_count)   if er else 0
        summary["active_days"]   = int(er.active_days)   if er else 0
        summary["docs_exported"] = int(er.docs_exported) if er else 0

        # Sessions created in range
        sc = await chat_db.execute(
            text("""
                SELECT COUNT(*) FROM sessions
                WHERE user_id = :uid AND created_at >= :start AND created_at <= :end
            """),
            {"uid": user_id, **range_params},
        )
        summary["sessions_created"] = int(sc.scalar() or 0)

        # Motion stats in range
        ms = await chat_db.execute(
            text("""
                SELECT
                    COUNT(m.id)                                                        AS motions_started,
                    COUNT(m.id) FILTER (WHERE m.status = 'completed')                 AS motions_completed,
                    AVG(
                        CASE WHEN m.status = 'completed' AND m.completed_at IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (m.completed_at - m.created_at)) END
                    )                                                                  AS avg_draft_time_seconds
                FROM motion_draft_logs m
                INNER JOIN sessions s ON s.id = m.session_id
                WHERE s.user_id = :uid AND m.created_at >= :start AND m.created_at <= :end
            """),
            {"uid": user_id, **range_params},
        )
        mr = ms.fetchone()
        started  = int(mr.motions_started)   if mr else 0
        done     = int(mr.motions_completed) if mr else 0
        avg_s    = float(mr.avg_draft_time_seconds) if mr and mr.avg_draft_time_seconds is not None else None
        summary["motions_started"]    = started
        summary["motions_completed"]  = done
        summary["draft_success_rate"] = round(done * 100.0 / started, 1) if started > 0 else None
        summary["avg_draft_time_sec"] = round(avg_s, 1) if avg_s is not None else None

        # ── Trend: motions per day ────────────────────────────────────────
        trend_m = await chat_db.execute(
            text("""
                SELECT DATE(m.created_at) AS day, COUNT(*) AS motions
                FROM motion_draft_logs m
                INNER JOIN sessions s ON s.id = m.session_id
                WHERE s.user_id = :uid
                  AND m.created_at >= :start AND m.created_at <= :end
                GROUP BY DATE(m.created_at)
            """),
            {"uid": user_id, **range_params},
        )
        motions_by_day: dict[date, int] = {r.day: int(r.motions) for r in trend_m}

        # Trend: active minutes per day (sum of duration_ms from activity logs)
        trend_a = await chat_db.execute(
            text("""
                SELECT
                    DATE(created_at) AS day,
                    ROUND(
                        (SUM(COALESCE((activity_metadata->>'duration_ms')::float, 0)) / 60000.0)::numeric,
                        1
                    ) AS active_minutes
                FROM user_activity_logs
                WHERE user_id = :uid
                  AND created_at >= :start AND created_at <= :end
                GROUP BY DATE(created_at)
            """),
            {"uid": user_id, **range_params},
        )
        minutes_by_day: dict[date, float] = {r.day: float(r.active_minutes) for r in trend_a}

        # Emit one row per calendar day that had any activity
        cur_d = dr.start.date()
        end_d = dr.end.date()
        while cur_d <= end_d:
            if cur_d in motions_by_day or cur_d in minutes_by_day:
                trend_rows.append((
                    cur_d.isoformat(),
                    motions_by_day.get(cur_d, 0),
                    minutes_by_day.get(cur_d, 0.0),
                ))
            cur_d += timedelta(days=1)

        # ── Top motion types in range ─────────────────────────────────────
        tm = await chat_db.execute(
            text("""
                SELECT
                    m.motion_type,
                    COUNT(*)                                              AS drafted,
                    COUNT(*) FILTER (WHERE m.status = 'completed')        AS completed
                FROM motion_draft_logs m
                INNER JOIN sessions s ON s.id = m.session_id
                WHERE s.user_id = :uid
                  AND m.created_at >= :start AND m.created_at <= :end
                GROUP BY m.motion_type
                ORDER BY drafted DESC
            """),
            {"uid": user_id, **range_params},
        )
        top_motion_rows = [(r.motion_type, int(r.drafted), int(r.completed)) for r in tm]

        # ── Recent sessions ───────────────────────────────────────────────
        sess_filters = ["s.user_id = :uid"]
        sess_params: dict = {"uid": user_id}

        if sessions_source:
            sess_filters.append("lp.source = :sess_source")
            sess_params["sess_source"] = sessions_source

        if sessions_status:
            sess_filters.append("lp.petition_status = :sess_status")
            sess_params["sess_status"] = sessions_status

        # Search by case number / debtor name via motion_draft_logs
        skip_sessions = False
        if sessions_search:
            ss_result = await chat_db.execute(
                text("""
                    SELECT DISTINCT session_id FROM motion_draft_logs
                    WHERE session_id IN (SELECT id FROM sessions WHERE user_id = :uid)
                      AND (
                          LOWER(COALESCE(case_number, '')) LIKE :q
                          OR LOWER(COALESCE(case_name, '')) LIKE :q
                      )
                """),
                {"uid": user_id, "q": f"%{sessions_search.lower()}%"},
            )
            found_ids = [r.session_id for r in ss_result]
            if found_ids:
                sess_filters.append("s.id = ANY(:ss_ids)")
                sess_params["ss_ids"] = found_ids
            else:
                skip_sessions = True

        if not skip_sessions:
            _SESS_SORT_COL = {
                "last_activity": "mdl.last_activity_at",
                "motions":       "mdl.motions_count",
                "source":        "lp.source",
                "status":        "lp.petition_status",
                "case":          "mdl.case_number",
            }
            sort_col  = _SESS_SORT_COL.get(sessions_sort_by, "mdl.last_activity_at")
            sort_dir_ = "DESC" if sessions_sort_dir == "desc" else "ASC"
            sess_offset = (sessions_page - 1) * sessions_page_size
            sess_where  = " AND ".join(sess_filters)

            sq = await chat_db.execute(
                text(f"""
                    WITH lp AS (
                        SELECT DISTINCT ON (session_id)
                            session_id, petition_status, source
                        FROM pdf_documents
                        ORDER BY session_id, uploaded_at DESC
                    ),
                    mdl AS (
                        SELECT
                            session_id,
                            MAX(COALESCE(case_number, ''))   AS case_number,
                            MAX(COALESCE(case_name, ''))     AS case_name,
                            COUNT(*)                         AS motions_count,
                            MAX(created_at)                  AS last_activity_at
                        FROM motion_draft_logs
                        GROUP BY session_id
                    )
                    SELECT
                        s.id                            AS session_id,
                        NULLIF(mdl.case_number, '')     AS case_number,
                        NULLIF(mdl.case_name, '')       AS debtor_name,
                        lp.source,
                        lp.petition_status,
                        COALESCE(mdl.motions_count, 0)  AS motions_count,
                        mdl.last_activity_at
                    FROM sessions s
                    LEFT JOIN lp  ON lp.session_id  = s.id
                    LEFT JOIN mdl ON mdl.session_id = s.id
                    WHERE {sess_where}
                    ORDER BY {sort_col} {sort_dir_} NULLS LAST
                    LIMIT :limit OFFSET :offset
                """),
                {**sess_params, "limit": sessions_page_size, "offset": sess_offset},
            )
            session_rows = sq.fetchall()

        # ── Recent activity ───────────────────────────────────────────────
        act_filters = [
            "user_id = :uid",
            "created_at >= :start",
            "created_at <= :end",
        ]
        act_params: dict = {"uid": user_id, **range_params}

        if activity_action:
            act_filters.append("action = :act_action")
            act_params["act_action"] = activity_action

        if activity_search:
            act_filters.append(
                "(action ILIKE :act_q"
                " OR activity_metadata->>'case_number' ILIKE :act_q"
                " OR activity_metadata->>'motion_type' ILIKE :act_q)"
            )
            act_params["act_q"] = f"%{activity_search}%"

        _ACT_SORT = {
            "occurred": "created_at",
            "action":   "action",
            "duration": "(activity_metadata->>'duration_ms')::float",
        }
        act_sort_col = _ACT_SORT.get(activity_sort_by, "created_at")
        act_sort_dir = "DESC" if activity_sort_dir == "desc" else "ASC"
        act_offset   = (activity_page - 1) * activity_page_size
        act_where    = " AND ".join(act_filters)

        aq = await chat_db.execute(
            text(f"""
                SELECT id, created_at, action, activity_metadata, session_id
                FROM user_activity_logs
                WHERE {act_where}
                ORDER BY {act_sort_col} {act_sort_dir} NULLS LAST
                LIMIT :limit OFFSET :offset
            """),
            {**act_params, "limit": activity_page_size, "offset": act_offset},
        )
        activity_rows = aq.fetchall()

    # ── Build workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Summary sheet (key-value layout) ─────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 42

    for key, val in [
        ("User ID",                user_id),
        ("Name",                   display_name),
        ("Email",                  u.email),
        ("Joined At (UTC)",        _fmt_dt(u.created_at)),
        ("Last Active At (UTC)",   _fmt_dt(summary.get("last_active_at"))),
        ("Login Count",            summary.get("login_count", 0)),
        ("Active Days",            summary.get("active_days", 0)),
        ("Sessions Created",       summary.get("sessions_created", 0)),
        ("Motions Started",        summary.get("motions_started", 0)),
        ("Motions Completed",      summary.get("motions_completed", 0)),
        ("Draft Success Rate (%)", summary.get("draft_success_rate")),
        ("Avg Draft Time (sec)",   summary.get("avg_draft_time_sec")),
        ("Documents Exported",     summary.get("docs_exported", 0)),
    ]:
        ws_sum.append([key, val if val is not None else ""])
        ws_sum.cell(row=ws_sum.max_row, column=1).font = _KEY_FONT

    # ── Trend_30d sheet ───────────────────────────────────────────────────
    ws_trend = wb.create_sheet("Trend_30d")
    _apply_header_row(ws_trend, ["Day", "Motions", "Active Minutes"], [14, 12, 16])
    for row in trend_rows:
        ws_trend.append(list(row))

    # ── Top_Motion_Types sheet ────────────────────────────────────────────
    ws_tm = wb.create_sheet("Top_Motion_Types")
    _apply_header_row(ws_tm, ["Motion Type", "Drafted", "Completed"], [38, 12, 14])
    for motion_type, drafted, completed in top_motion_rows:
        display = _MOTION_DISPLAY_NAMES.get(motion_type, motion_type)
        ws_tm.append([display, drafted, completed])

    # ── Recent_Sessions sheet ─────────────────────────────────────────────
    ws_sess = wb.create_sheet("Recent_Sessions")
    _apply_header_row(ws_sess, [
        "Session ID", "Case Number", "Debtor Name",
        "District", "Source", "Petition Status",
        "Motions Count", "Last Activity At (UTC)",
    ], [38, 20, 30, 16, 14, 20, 16, 24])
    for r in session_rows:
        ws_sess.append([
            str(r.session_id),
            r.case_number or "",
            r.debtor_name or "",
            "",                       # district: not stored in current schema
            r.source or "",
            r.petition_status or "",
            int(r.motions_count),
            _fmt_dt(r.last_activity_at),
        ])

    # ── Recent_Activity sheet ─────────────────────────────────────────────
    ws_act = wb.create_sheet("Recent_Activity")
    _apply_header_row(ws_act, [
        "ID", "Occurred At (UTC)", "Action", "Detail",
        "Entity ID", "Status", "Duration (ms)",
    ], [38, 22, 30, 42, 38, 14, 14])
    for r in activity_rows:
        meta    = r.activity_metadata or {}
        detail  = _build_activity_detail(r.action, meta) or ""
        e_id    = meta.get("case_number") or (str(r.session_id) if r.session_id else "")
        raw_dur = meta.get("duration_ms")
        try:
            dur = float(raw_dur) if raw_dur is not None else None
        except (TypeError, ValueError):
            dur = None
        ws_act.append([
            str(r.id),
            _fmt_dt(r.created_at),
            _ACTION_LABELS.get(r.action, r.action),
            detail,
            e_id,
            "",      # status: batch lookup against motion_draft_logs omitted for export
            dur,
        ])

    # ── Meta sheet ────────────────────────────────────────────────────────
    applied_filters: dict = {}
    if sessions_source:
        applied_filters["sessions_source"] = sessions_source
    if sessions_status:
        applied_filters["sessions_status"] = sessions_status
    if activity_action:
        applied_filters["activity_action"] = activity_action

    _write_meta_sheet(wb, [
        ("Export Generated At (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Export Type", "single_user"),
        ("Range Preset", dr.range),
        ("Range Start (UTC)", dr.start.strftime("%Y-%m-%d %H:%M UTC")),
        ("Range End (UTC)", dr.end.strftime("%Y-%m-%d %H:%M UTC")),
        ("Applied Search", sessions_search or activity_search or ""),
        ("Applied Filters (JSON)", json.dumps(applied_filters)),
        ("Applied Sort (JSON)", json.dumps({
            "sessions_sort_by":  sessions_sort_by,
            "sessions_sort_dir": sessions_sort_dir,
            "activity_sort_by":  activity_sort_by,
            "activity_sort_dir": activity_sort_dir,
        })),
        ("Total Rows Exported", str(len(session_rows) + len(activity_rows))),
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    date_tag   = datetime.now(timezone.utc).strftime("%Y%m%d")
    safe_name  = display_name.replace(" ", "_")[:20]
    filename   = f"user_{safe_name}_{dr.range}_{date_tag}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
